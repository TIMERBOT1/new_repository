import csv
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pdfkit
from jinja2 import FileSystemLoader, Environment

class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.сsv_reader(file_name)

    def сsv_reader(file_name):
        file_csv = open(file_name, encoding='utf_8_sig')
        reader_csv = csv.reader(file_csv)
        listData = []
        for x in reader_csv:
            listData.append(x)
        if len(listData) == 0:
            print('Пустой файл')
            exit(0)
        if len(listData) == 1:
            print("Нет данных")
            exit(0)
        columns = listData[0]
        columns = dict(zip(columns, list(range(len(columns)))))
        data = listData[1:]
        class_data = []
        for x in data:
            g = [x[columns['name']], x[columns['salary_from']], x[columns['salary_to']], x[columns["salary_currency"]], x[columns['area_name']], x[columns['published_at']]]
            if len(columns) == len(x) and not x.__contains__(''):
                class_data.append(Vacancy(g))
        return class_data


class Salary:

    def __init__(self, salary_from, salary_to, salary_currency):
        currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                           "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
        self.salary = int((float(salary_to) + float(salary_from))/2 * currency_to_rub[salary_currency])
        self.salary_currency = salary_currency

class Vacancy:

    def __init__(self, vacanlist):
        self.name = vacanlist[0]
        self.salary = Salary(vacanlist[1], vacanlist[2], vacanlist[3])
        self.area_name = vacanlist[4]
        self.published_at = int(vacanlist[5][:4])


class InputConnect:

    def __init__(self):
        self.params = InputConnect.get_params(self)
        self.filename = self.params[0]
        self.name = self.params[1]


    def get_params(self):
        filename = input('Введите название файла: ')
        name = input('Введите название профессии: ')
        return [filename, name]

    def statistics_for_years(self, name):
        list1 = a.vacancies_objects
        def cities_statistic():
            if city_salary.__contains__(i.area_name):
                city_salary[i.area_name] = city_salary[i.area_name] + i.salary.salary
                city_count[i.area_name] = city_count[i.area_name] + 1
            else:
                city_salary[i.area_name] = i.salary.salary
                city_count[i.area_name] = 1
        salary = {}
        count = {}
        salary_by_vacancy = {}
        count_by_vacancy = {}
        city_salary = {}
        city_count = {}
        for i in list1:
            if salary.__contains__(i.published_at):
                salary[i.published_at] = i.salary.salary + salary[i.published_at]
                count[i.published_at] = count[i.published_at] + 1
                if i.name.__contains__(name):
                    salary_by_vacancy[i.published_at] = i.salary.salary + salary_by_vacancy[i.published_at]
                    count_by_vacancy[i.published_at] = count_by_vacancy[i.published_at] + 1
                cities_statistic()
            else:
                salary[i.published_at] = i.salary.salary
                count[i.published_at] = 1
                if i.name.__contains__(name):
                    salary_by_vacancy[i.published_at] = i.salary.salary
                    count_by_vacancy[i.published_at] = 1
                else:
                    salary_by_vacancy[i.published_at] = 0
                    count_by_vacancy[i.published_at] = 0
                cities_statistic()
        for key, value in salary.items():
            salary[key] = int(value/count[key])

        for key1, value1 in salary_by_vacancy.items():
            salary_by_vacancy[key1] = int(value1 / count_by_vacancy[key1]) if value1 != 0 else 0

        for key, value in list(city_salary.items()):
            city_salary[key] = int(value / city_count[key])
            if len(list1) * 0.01 > city_count[key]:
                del city_salary[key]
                del city_count[key]
                continue
            city_count[key] = round(city_count[key] / len(list1), 4)

        city_salary = dict(sorted(city_salary.items(), key=lambda x: x[1], reverse=True)[:10])
        city_count = dict(sorted(city_count.items(), key=lambda x: x[1], reverse=True)[:10])
        return salary, count, city_salary, city_count, salary_by_vacancy, count_by_vacancy


class Report:

    def generate_pdf(self):
        options = {'enable-local-file-access': None}
        image = 'report.png'
        name = 'Программист'
        book = load_workbook('report.xlsx')
        sheet_1 = book.active
        sheet_2 = book['Статистика по городам']
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'name': name, 'sheet_1': sheet_1, 'sheet_2': sheet_2})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'out.pdf', configuration=config, options=options)

    def generate_image(self, statistic):
        salary, count, salary_city, city_count, vacancy_salary, vacancy_count = statistic

        def bar_chart(first_dict, second_dict, label1, label2):
            x = np.arange(len(first_dict.keys()))
            ax.set_title('Уровень зарплат по годам')
            ax.bar(x - 0.15, first_dict.values(), 0.3, label=label1)
            ax.bar(x + 0.15, second_dict.values(), 0.3, label=label2)
            ax.set_xticks(x, first_dict.keys(), rotation='vertical')
            ax.set_yticks([20000, 40000, 60000, 80000, 100000, 120000])
            for label in (ax.get_xticklabels() + ax.get_yticklabels()):
                label.set_fontsize(8)
            ax.grid(axis='y')
            ax.legend(fontsize=8)

        fig = plt.figure()
        ax = fig.add_subplot(221)
        bar_chart(salary, vacancy_salary, 'средняя з/п', f'з/п {m.name.lower()}')

        ax = fig.add_subplot(222)
        bar_chart(count, vacancy_count, 'Количество вакансий', f'Количество вакансий\n{m.name.lower()}')
        ax.set_title('Количество вакансий по годам')

        ax = fig.add_subplot(223)
        y = np.arange(len(list(salary_city.keys())[::-1]))
        ax.barh(y, list(salary_city.values())[::-1])
        ax.grid(axis='x')
        ax.set_yticks(y, list(salary_city.keys())[::-1])
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(8)
        ax.set_title('Уровень зарплат по городам')

        ax = fig.add_subplot(224)
        count_dict = {'Другие': 1 - sum(city_count.values())}
        count_dict.update(city_count)
        ax.pie(count_dict.values(), labels=count_dict.keys(), radius=1.1, textprops={"fontsize": 6})
        ax.set_title('Доля вакансий по городам')

        plt.tight_layout()
        plt.savefig('report.png')

    def generate_exel(self, statistic, name):

        def format_to_precent(e):
            e = round(e*100, 2)
            return f'{e}%'

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wb = Workbook()
        sheet1 = wb.create_sheet("Статистика по годам")
        sheet2 = wb.create_sheet("Статистика по городам")
        heads = ["Год", "Средняя зарплата", f"Средняя зарплата - {m.name}", "Количество вакансий",
                 f"Количество вакансий - {m.name}"]
        for i, head in enumerate(heads, start=1):
            sheet1.cell(row=1, column=i, value=head).font = Font(bold=True)
        s1 = 1
        for year, value in statistic[0].items():
            sheet1.append([year, value, statistic[4][year], statistic[1][year], statistic[5][year]])
            s1 += 1
        set_border(sheet1, f'A1:E{s1}')
        heads2 = {1: "Город", 2: "Уровень зарплат", 4: "Город", 5: "Доля вакансий"}
        for i, city in heads2.items():
            sheet2.cell(row=1, column=i, value=city).font = Font(bold=True)
        h = 2
        for city, value in statistic[2].items():
            sheet2.cell(row=h, column=1, value=city)
            sheet2.cell(row=h, column=2, value=value)
            h += 1
        set_border(sheet2, f'A1:B{h-1}')
        set_border(sheet2, f'D1:E{h-1}')
        h = 2
        for city, value in statistic[3].items():
            sheet2.cell(row=h, column=4, value=city)
            sheet2.cell(row=h, column=5, value=format_to_precent(value))
            h += 1
        for cel in sheet2['E']:
            cel.alignment = Alignment(horizontal='right')
        sheet1.column_dimensions['A'].width = 4
        sheet1.column_dimensions['B'].width = 16
        sheet1.column_dimensions['C'].width = 18 + len(name)
        sheet1.column_dimensions['D'].width = 19
        sheet1.column_dimensions['E'].width = 20 + len(name)

        sheet2.column_dimensions['A'].width = 17
        sheet2.column_dimensions['B'].width = 16
        sheet2.column_dimensions['C'].width = 4
        sheet2.column_dimensions['D'].width = 17
        sheet2.column_dimensions['E'].width = 14
        del wb['Sheet']
        wb.save('report.xlsx')


def get_report():
    m = InputConnect()
    a = DataSet(m.filename)
    statistic = m.statistics_for_years(m.name)
    salary, count, salary_city, city_count, vacancy_salary, vacancy_count = statistic
    print(f'Динамика уровня зарплат по годам: {salary}\nДинамика количества вакансий по годам: {count}\n'
          f'Динамика уровня зарплат по годам для выбранной профессии: {vacancy_salary}\nДинамика количества вакансий по годам для выбранной профессии: {vacancy_count}\n'
          f'Уровень зарплат по городам (в порядке убывания): {salary_city}\nДоля вакансий по городам (в порядке убывания): {city_count}')

    Report().generate_image(statistic)
    Report().generate_exel(statistic, m.name)
    Report().generate_pdf()


